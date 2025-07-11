#!/usr/bin/env python
# -*- coding: utf-8 -*-

from flask import Blueprint, request, send_file, current_app as app
import pymysql
import requests
import json
import time
import uuid
import concurrent.futures
import threading
from datetime import datetime
from io import BytesIO
import traceback
from configs import config, format
#from utils.auth import get_current_user
from dbutils.pooled_db import PooledDB
#import jsonpath
import openpyxl
from openpyxl.styles import Font
from app import celery, global_logger
from utils import metrics

# 创建蓝图
testexec = Blueprint('testexec', __name__)

#获取数据库连接池
mysql_pool=PooledDB(
    creator=pymysql,
    maxconnections=6,
    mincached=2,
    maxcached=5,
    maxshared=3,
    blocking=True,
    maxusage=None,
    setsession=[],
    ping=0,
    host=config.MYSQL_HOST,
    port=config.MYSQL_PORT,
    user=config.MYSQL_USER,
    password=config.MYSQL_PASSWORD,
    database=config.MYSQL_DATABASE,
    cursorclass=pymysql.cursors.DictCursor
)


@testexec.route('/api/testexec/execute', methods=['POST'])
def execute_testcase():
    """
    执行单个测试用例
    """

    global_logger.info('访问执行单个测试用例API')
    data = request.json
    global_logger.info(f'请求参数: {json.dumps(data)}')

    # 检查必要参数
    if 'testcase_id' not in data or not data['testcase_id']:
        global_logger.error('缺少必要参数: testcase_id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: testcase_id"
        return response

        # 获取当前用户
    current_user = get_current_user()
    if not current_user:
        global_logger.error('获取当前用户失败，用户未登录或登录已过期')
        response = format.resp_format_failed.copy()
        response["message"] = "用户未登录或登录已过期"
        response["code"] = 50008
        return response

    testcase_id = data['testcase_id']
    environment = data.get('environment', 'test')  # 默认使用测试环境
    variables = data.get('variables', {})

    global_logger.info(f'开始执行测试用例，ID: {testcase_id}, 环境: {environment}')

    try:
        result = execute_single_testcase(testcase_id, environment, variables, current_user)
        global_logger.info(f"测试用例执行完成，结果ID: {result['result_id']}, 是否成功: {result.get('is_success', False)}")

        response = format.resp_format_success.copy()
        response["message"] = "测试用例执行成功"
        response["data"] = result
        return response
    except Exception as e:
        global_logger.error(f"执行测试用例异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response

@testexec.route('/api/testexec/batch_execute', methods=['POST'])
def batch_execute_testcase():
    """
    批量执行测试用例
    """
    global_logger.info('访问批量执行测试用例API')
    data = request.json
    global_logger.info(f'请求参数: {json.dumps(data)}')

    # 检查必要参数
    required_fields = ['testcase_ids', 'app_id', 'name']
    for field in required_fields:
        if field not in data or not data[field]:
            global_logger.error(f'缺少必要参数: {field}')
            response = format.resp_format_failed.copy()
            response["message"] = f"缺少必要参数: {field}"
            return response

            # 获取当前用户
    current_user = get_current_user()
    if not current_user:
        global_logger.error('获取当前用户失败，用户未登录或登录已过期')
        response = format.resp_format_failed.copy()
        response["message"] = "用户未登录或登录已过期"
        response["code"] = 50008
        return response

    testcase_ids = data['testcase_ids']
    app_id = data['app_id']
    test_request_id = data.get('test_request_id')
    name = data['name']
    environment = data.get('environment', 'test')
    async_mode = data.get('async_mode', True)

    #  关键：检查是否传入了batch_id
    existing_batch_id = data.get('batch_id')

    global_logger.info(
        f'开始批量执行测试用例，数量: {len(testcase_ids)}, 应用ID: {app_id}, 环境: {environment}, 异步模式: {async_mode}, 现有批次ID: {existing_batch_id}')

    try:
        #  批次处理逻辑
        if existing_batch_id:
            # 使用现有批次
            batch_id = existing_batch_id
            global_logger.info(f'使用现有批次ID: {batch_id}')

            #  验证批次是否存在
            if not verify_batch_exists(batch_id):
                global_logger.error(f'批次不存在: {batch_id}')
                response = format.resp_format_failed.copy()
                response["message"] = f"批次不存在: {batch_id}"
                return response

                #  更新批次状态为执行中
            update_batch_status_to_running(batch_id)

        else:
            # 创建新批次（传统流程）
            batch_id = str(uuid.uuid4()).replace('-', '')
            global_logger.info(f'生成新批次ID: {batch_id}')
            #  传递ai_generated参数
            ai_generated = 1 if data.get('ai_generated') else 0
            create_test_batch(batch_id, app_id, test_request_id, name, len(testcase_ids), current_user, ai_generated)
            global_logger.info(f"创建测试批次成功，ID: {batch_id}, 名称: {name}, 测试用例数量: {len(testcase_ids)}")
            # 异步执行测试用例
        if async_mode:
            global_logger.info(f'启动异步执行线程，批次ID: {batch_id}')
            execute_batch_testcases_async.delay(batch_id, testcase_ids, environment, current_user)
            global_logger.info(f"已提交异步任务，批次ID: {batch_id}")
            response = format.resp_format_success.copy()
            response["message"] = "测试批次已创建，测试用例正在异步执行中" if not existing_batch_id else "测试批次正在异步执行中"
            response["data"] = {
                "batch_id": batch_id,
                "total_cases": len(testcase_ids)
            }
            return response
        else:
            global_logger.info(f'开始同步执行测试用例，批次ID: {batch_id}')
            # 同步执行测试用例
            results = execute_batch_testcases_sync(batch_id, testcase_ids, environment, current_user)

            # 计算统计数据
            passed_cases = sum(1 for r in results if r.get('is_success'))
            failed_cases = sum(1 for r in results if not r.get('is_success'))

            # 更新批次统计数据
            update_batch_progress(batch_id, passed_cases, failed_cases)

            # 更新批次状态
            update_batch_status(batch_id)

            passed_cases = sum(1 for r in results if r.get('is_success'))
            failed_cases = sum(1 for r in results if not r.get('is_success'))
            global_logger.info(f'测试批次执行完成，批次ID: {batch_id}, 通过: {passed_cases}, 失败: {failed_cases}')

            response = format.resp_format_success.copy()
            response["message"] = "测试批次执行完成"
            response["data"] = {
                "batch_id": batch_id,
                "total_cases": len(testcase_ids),
                "passed_cases": passed_cases,
                "failed_cases": failed_cases
            }
            return response
    except Exception as e:
        global_logger.error(f"批量执行测试用例异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response

@testexec.route('/api/testexec/batch_list', methods=['GET'])
def get_batch_list():
    """
    获取测试批次列表
    """
    global_logger.info('访问获取测试批次列表API')

    # 获取请求参数
    app_id = request.args.get('app_id')
    test_request_id = request.args.get('test_request_id')
    page = int(request.args.get('page', 1))
    size = int(request.args.get('size', 10))

    global_logger.info(f'请求参数: app_id={app_id}, test_request_id={test_request_id}, page={page}, size={size}')

    # 验证参数
    if not app_id:
        global_logger.error('缺少必要参数: app_id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: app_id"
        return response

        # 计算分页
    offset = (page - 1) * size

    # 查询数据库
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 构建查询条件
        where_clause = "WHERE app_id = %s"
        params = [app_id]

        if test_request_id:
            where_clause += " AND test_request_id = %s"
            params.append(test_request_id)

            # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM api_test_batch {where_clause}"
        global_logger.info(f'执行SQL: {count_sql}, 参数: {params}')
        cursor.execute(count_sql, params)
        total = cursor.fetchone()['total']

        # 查询列表
        sql = f"""  
            SELECT id, name, total_cases, passed_cases, failed_cases, status,   
                   executor_name, create_time  
            FROM api_test_batch  
            {where_clause}  
            ORDER BY create_time DESC  
            LIMIT %s, %s  
        """
        params.extend([offset, size])
        global_logger.info(f'执行SQL: {sql}, 参数: {params}')

        cursor.execute(sql, params)
        batch_list = cursor.fetchall()

        global_logger.info(f"获取测试批次列表成功，总数: {total}, 当前页数量: {len(batch_list)}")

        response = format.resp_format_success.copy()
        response["message"] = "获取测试批次列表成功"
        response["data"] = batch_list
        response["total"] = total
        return response
    except Exception as e:
        global_logger.error(f"获取测试批次列表异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response
    finally:
        cursor.close()
        conn.close()


@testexec.route('/api/testexec/batch_detail', methods=['GET'])
def get_batch_detail():
    """
    获取测试批次详情
    """
    global_logger.info('访问获取测试批次详情API')

    # 获取请求参数
    batch_id = request.args.get('id')
    global_logger.info(f'请求参数: id={batch_id}')

    # 验证参数
    if not batch_id:
        global_logger.error('缺少必要参数: id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: id"
        return response

        # 查询数据库
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        sql = """  
            SELECT id, name, app_id, test_request_id, total_cases, passed_cases,   
                   failed_cases, status, trigger_type, executor_id, executor_name,   
                   create_time, end_time  
            FROM api_test_batch  
            WHERE id = %s  
        """
        global_logger.info(f'执行SQL: {sql}, 参数: [{batch_id}]')
        cursor.execute(sql, [batch_id])
        batch = cursor.fetchone()

        if not batch:
            global_logger.warning(f"未找到测试批次，ID: {batch_id}")
            response = format.resp_format_failed.copy()
            response["message"] = "未找到测试批次"
            response["code"] = 40004
            return response

        global_logger.info(f"获取测试批次详情成功，ID: {batch_id}, 名称: {batch['name']}")

        response = format.resp_format_success.copy()
        response["message"] = "获取测试批次详情成功"
        response["data"] = batch
        return response
    except Exception as e:
        global_logger.error(f"获取测试批次详情异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response
    finally:
        cursor.close()
        conn.close()


@testexec.route('/api/testexec/result_list', methods=['GET'])
def get_result_list():
    """
    获取测试结果列表
    """
    global_logger.info('访问获取测试结果列表API')

    # 获取请求参数
    batch_id = request.args.get('batch_id')
    is_success = request.args.get('is_success')
    page = int(request.args.get('page', 1))
    size = int(request.args.get('size', 10))

    global_logger.info(f'请求参数: batch_id={batch_id}, is_success={is_success}, page={page}, size={size}')

    # 验证参数
    if not batch_id:
        global_logger.error('缺少必要参数: batch_id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: batch_id"
        return response

        # 计算分页
    offset = (page - 1) * size

    # 查询数据库
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 构建查询条件
        where_clause = "WHERE batch_id = %s"
        params = [batch_id]

        if is_success is not None:
            where_clause += " AND is_success = %s"
            params.append(int(is_success))

            # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM api_result {where_clause}"
        global_logger.info(f'执行SQL: {count_sql}, 参数: {params}')
        cursor.execute(count_sql, params)
        total = cursor.fetchone()['total']

        # 查询列表
        sql = f"""  
            SELECT id, testcase_id, request_url, request_method, response_status,   
                   is_success, execution_time, execute_time  
            FROM api_result  
            {where_clause}  
            ORDER BY execute_time DESC  
            LIMIT %s, %s  
        """
        params.extend([offset, size])
        global_logger.info(f'执行SQL: {sql}, 参数: {params}')

        cursor.execute(sql, params)
        result_list = cursor.fetchall()

        global_logger.info(f"获取测试结果列表成功，总数: {total}, 当前页数量: {len(result_list)}")

        response = format.resp_format_success.copy()
        response["message"] = "获取测试结果列表成功"
        response["data"] = result_list
        response["total"] = total
        return response
    except Exception as e:
        global_logger.error(f"获取测试结果列表异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response
    finally:
        cursor.close()
        conn.close()


@testexec.route('/api/testexec/result_detail', methods=['GET'])
def get_result_detail():
    """
    获取测试结果详情
    """
    global_logger.info('访问获取测试结果详情API')

    # 获取请求参数
    result_id = request.args.get('id')
    global_logger.info(f'请求参数: id={result_id}')

    # 验证参数
    if not result_id:
        global_logger.error('缺少必要参数: id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: id"
        return response

        # 查询数据库
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        sql = """  
            SELECT id, testcase_id, interface_id, app_id, batch_id, request_url,   
                   request_method, request_headers, request_body, response_status,   
                   response_headers, response_body, assertion_results, is_success,   
                   error_message, execution_time, executor_id, executor_name, execute_time  
            FROM api_result  
            WHERE id = %s  
        """
        global_logger.info(f'执行SQL: {sql}, 参数: [{result_id}]')
        cursor.execute(sql, [result_id])
        result = cursor.fetchone()

        if not result:
            global_logger.warning(f"未找到测试结果，ID: {result_id}")
            response = format.resp_format_failed.copy()
            response["message"] = "未找到测试结果"
            response["code"] = 40004
            return response

            # 处理JSON字段
        global_logger.info(f"开始处理测试结果的JSON字段，ID: {result_id}")
        if result['request_headers']:
            result['request_headers'] = json.loads(result['request_headers'])
        if result['request_body']:
            result['request_body'] = json.loads(result['request_body'])
        if result['response_headers']:
            result['response_headers'] = json.loads(result['response_headers'])
        if result['response_body']:
            try:
                result['response_body'] = json.loads(result['response_body'])
            except:
                global_logger.warning(f"响应体解析JSON失败，保持原始字符串: {result_id}")
        if result['assertion_results']:
            result['assertion_results'] = json.loads(result['assertion_results'])

        global_logger.info(f"获取测试结果详情成功，ID: {result_id}, 是否成功: {result['is_success']}")

        response = format.resp_format_success.copy()
        response["message"] = "获取测试结果详情成功"
        response["data"] = result
        return response
    except Exception as e:
        global_logger.error(f"获取测试结果详情异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response
    finally:
        cursor.close()
        conn.close()


@testexec.route('/api/testexec/export_report', methods=['GET'])
def export_test_report():
    """
    导出测试报告
    """
    global_logger.info('访问导出测试报告API')

    # 获取请求参数
    batch_id = request.args.get('batch_id')
    report_type = request.args.get('type', 'excel')  # 默认导出Excel格式

    global_logger.info(f'请求参数: batch_id={batch_id}, type={report_type}')

    # 验证参数
    if not batch_id:
        global_logger.error('缺少必要参数: batch_id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: batch_id"
        return response

    try:
        if report_type == 'excel':
            global_logger.info(f"开始生成Excel测试报告，批次ID: {batch_id}")
            # 生成Excel报告
            excel_data = generate_excel_report(batch_id)

            # 获取批次信息用于文件命名
            conn = mysql_pool.connection()
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM api_test_batch WHERE id = %s", [batch_id])
            batch_name = cursor.fetchone()['name']
            cursor.close()
            conn.close()

            filename = f"测试报告_{batch_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

            global_logger.info(f"生成Excel测试报告成功，文件名: {filename}")

            # 返回Excel文件
            return send_file(
                BytesIO(excel_data),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                attachment_filename=filename
            )
        elif report_type == 'pdf':
            global_logger.info(f"开始生成PDF测试报告，批次ID: {batch_id}")
            # 生成PDF报告
            pdf_data = generate_pdf_report(batch_id)

            # 获取批次信息用于文件命名
            conn = mysql_pool.connection()
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM api_test_batch WHERE id = %s", [batch_id])
            batch_name = cursor.fetchone()['name']
            cursor.close()
            conn.close()

            filename = f"测试报告_{batch_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"

            global_logger.info(f"生成PDF测试报告成功，文件名: {filename}")

            # 返回PDF文件
            return send_file(
                BytesIO(pdf_data),
                mimetype='application/pdf',
                as_attachment=True,
                attachment_filename=filename
            )
        else:
            global_logger.error(f"不支持的报告类型: {report_type}")
            response = format.resp_format_failed.copy()
            response["message"] = f"不支持的报告类型: {report_type}"
            return response
    except Exception as e:
        global_logger.error(f"导出测试报告异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response


@testexec.route('/api/testexec/dashboard', methods=['GET'])
def get_test_dashboard():
    """
    获取测试执行统计数据
    """
    global_logger.info('访问测试执行统计数据API')

    # 获取请求参数
    app_id = request.args.get('app_id')
    period = request.args.get('period', 'week')  # 默认统计周期为一周

    global_logger.info(f'请求参数: app_id={app_id}, period={period}')

    # 验证参数
    if not app_id:
        global_logger.error('缺少必要参数: app_id')
        response = format.resp_format_failed.copy()
        response["message"] = "缺少必要参数: app_id"
        return response

        # 查询数据库
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 确定时间范围
        current_time = int(time.time())
        if period == 'day':
            time_range = current_time - 86400  # 一天
        elif period == 'week':
            time_range = current_time - 604800  # 一周
        elif period == 'month':
            time_range = current_time - 2592000  # 30天
        else:
            time_range = current_time - 604800  # 默认一周

        global_logger.info(f"统计时间范围: {datetime.fromtimestamp(time_range)} 至 {datetime.fromtimestamp(current_time)}")

        # 查询测试批次统计
        batch_sql = """  
            SELECT COUNT(*) as total_batches,   
                   SUM(passed_cases) as total_passed,   
                   SUM(failed_cases) as total_failed,  
                   SUM(total_cases) as total_cases  
            FROM api_test_batch  
            WHERE app_id = %s AND create_time > %s  
        """
        global_logger.info(f'执行SQL: {batch_sql}, 参数: [{app_id}, {time_range}]')
        cursor.execute(batch_sql, [app_id, time_range])
        batch_stats = cursor.fetchone()

        # 查询每日执行测试用例数量
        daily_sql = """  
            SELECT FROM_UNIXTIME(create_time, '%%Y-%%m-%%d') as date,   
                   COUNT(*) as total,  
                   SUM(CASE WHEN status = 2 THEN 1 ELSE 0 END) as completed,  
                   SUM(passed_cases) as passed,  
                   SUM(failed_cases) as failed  
            FROM api_test_batch  
            WHERE app_id = %s AND create_time > %s  
            GROUP BY date  
            ORDER BY date  
        """
        global_logger.info(f'执行SQL: {daily_sql}, 参数: [{app_id}, {time_range}]')
        cursor.execute(daily_sql, [app_id, time_range])
        daily_stats = cursor.fetchall()

        # 查询接口覆盖率
        coverage_sql = """  
            SELECT COUNT(DISTINCT i.id) as total_interfaces,  
                   COUNT(DISTINCT CASE WHEN r.id IS NOT NULL THEN i.id END) as covered_interfaces  
            FROM api_interface i  
            LEFT JOIN (  
                SELECT DISTINCT interface_id   
                FROM api_result   
                WHERE app_id = %s AND execute_time > %s  
            ) r ON i.id = r.interface_id  
            WHERE i.app_id = %s  
        """
        global_logger.info(f'执行SQL: {coverage_sql}, 参数: [{app_id}, {time_range}, {app_id}]')
        cursor.execute(coverage_sql, [app_id, time_range, app_id])
        coverage = cursor.fetchone()

        global_logger.info(f"获取测试执行统计数据成功，应用ID: {app_id}")

        # 构建响应数据
        dashboard_data = {
            'batch_stats': {
                'total_batches': batch_stats['total_batches'] or 0,
                'total_cases': batch_stats['total_cases'] or 0,
                'total_passed': batch_stats['total_passed'] or 0,
                'total_failed': batch_stats['total_failed'] or 0,
                'success_rate': round((batch_stats['total_passed'] or 0) * 100 / (batch_stats['total_cases'] or 1), 2)
            },
            'daily_stats': daily_stats,
            'interface_coverage': {
                'total_interfaces': coverage['total_interfaces'] or 0,
                'covered_interfaces': coverage['covered_interfaces'] or 0,
                'coverage_rate': round(
                    (coverage['covered_interfaces'] or 0) * 100 / (coverage['total_interfaces'] or 1), 2)
            }
        }

        response = format.resp_format_success.copy()
        response["message"] = "获取测试执行统计数据成功"
        response["data"] = dashboard_data
        return response
    except Exception as e:
        global_logger.error(f"获取测试执行统计数据异常: {str(e)}")
        global_logger.error(traceback.format_exc())

        response = format.resp_format_failed.copy()
        response["message"] = f"系统异常: {str(e)}"
        return response
    finally:
        cursor.close()
        conn.close()


def execute_single_testcase(testcase_id, environment='test', variables={}, current_user=None):
    """
    执行单个测试用例

    参数:
        testcase_id: 测试用例ID
        environment: 执行环境，默认为test
        variables: 环境变量，用于替换请求中的变量
        current_user: 当前用户信息

    返回:
        包含执行结果的字典
    """
    global_logger.info(f'开始执行单个测试用例, ID: {testcase_id}, 环境: {environment}')

    start_time=time.time()

    # 生成结果ID
    result_id = str(uuid.uuid4()).replace('-', '')
    global_logger.info(f'生成测试结果ID: {result_id}')

    # 获取测试用例信息
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 查询测试用例
        sql = """  
            SELECT id, interface_id, app_id, name, priority, request_url, request_method,   
                   request_headers, request_params, expected_status, assertions,   
                   pre_script, post_script  
            FROM api_testcase  
            WHERE id = %s  
        """
        global_logger.info(f'执行SQL: {sql}, 参数: [{testcase_id}]')
        cursor.execute(sql, [testcase_id])
        testcase = cursor.fetchone()

        if not testcase:
            global_logger.error(f'未找到测试用例, ID: {testcase_id}')
            raise Exception(f"未找到测试用例: {testcase_id}")

        global_logger.info(f'获取测试用例信息成功, 名称: {testcase["name"]}')

        # 获取环境配置
        env_sql = """  
            SELECT base_url, headers, global_variables  
            FROM api_environment  
            WHERE app_id = %s AND env = %s  
        """
        global_logger.info(f'执行SQL: {env_sql}, 参数: [{testcase["app_id"]}, {environment}]')
        cursor.execute(env_sql, [testcase["app_id"], environment])
        env_config = cursor.fetchone()

        if not env_config:
            global_logger.error(f'未找到环境配置, 应用ID: {testcase["app_id"]}, 环境: {environment}')
            raise Exception(f"未找到环境配置: app_id={testcase['app_id']}, environment={environment}")

        global_logger.info(f'获取环境配置成功, 基础URL: {env_config["base_url"]}')

        # 准备请求参数
        request_url = testcase['request_url']

        # 如果URL不是以http开头，则拼接基础URL
        if not request_url.startswith(('http://', 'https://')):
            request_url = env_config['base_url'].rstrip('/') + '/' + request_url.lstrip('/')

        request_method = testcase['request_method'].upper()

        # 处理请求头
        request_headers = {}
        if env_config['headers']:
            try:
                env_headers = json.loads(env_config['headers'])
                request_headers.update(env_headers)
            except Exception as e:
                global_logger.warning(f'解析环境请求头异常: {str(e)}')

        if testcase['request_headers']:
            try:
                case_headers = json.loads(testcase['request_headers'])
                request_headers.update(case_headers)
            except Exception as e:
                global_logger.warning(f'解析测试用例请求头异常: {str(e)}')

                # 处理请求参数
        request_params = {}
        request_data = None
        request_json = None

        global_logger.info(f'测试用例ID: {testcase_id}')
        global_logger.info(f'request_params原始值: {repr(testcase["request_params"])}')
        global_logger.info(f'request_params是否为真: {bool(testcase["request_params"])}')

        if testcase['request_params']:
            try:
                params_data = json.loads(testcase['request_params'])
                global_logger.info(f'解析params_data成功: {params_data}')

                # 处理GET请求参数
                if request_method == 'GET':
                    request_params = params_data
                    # 处理POST/PUT/DELETE请求体
                elif request_method in ['POST', 'PUT', 'DELETE', 'PATCH']:
                    request_json = params_data
                    if 'Content-Type' not in request_headers:
                        request_headers['Content-Type'] = 'application/json'
            except Exception as e:
                global_logger.error(f'解析测试用例请求参数异常: {str(e)}')

        global_logger.info(f'最终处理结果 - request_params: {request_params}')
        global_logger.info(f'最终处理结果 - request_json: {request_json}')
        global_logger.info(f'最终处理结果')

        #记录当前时间
        http_request_start_time = time.time()

        global_logger.info(f'准备执行HTTP请求, URL: {request_url}, 方法: {request_method}')
        global_logger.info(f'请求头: {request_headers}')
        global_logger.info(f'请求参数 (params): {request_params}')
        global_logger.info(f'请求数据 (data): {request_data}')
        global_logger.info(f'请求JSON (json): {request_json}')
        # 执行前置脚本
        if testcase['pre_script']:
            global_logger.info(f'执行前置脚本, 测试用例ID: {testcase_id}')
            try:
                # 这里可以实现执行前置脚本的逻辑
                pass
            except Exception as e:
                global_logger.error(f'执行前置脚本异常: {str(e)}')

                # 记录开始时间
        start_time = time.time()

        # 发送HTTP请求
        try:
            global_logger.info(f'准备发送请求 - 方法: {request_method}, URL: {request_url}')
            global_logger.info(f'请求头: {request_headers}')
            global_logger.info(f'请求JSON: {request_json}')
            global_logger.info(f'请求JSON类型: {type(request_json)}')

            if request_method == 'GET':
                response = requests.get(
                    url=request_url,
                    headers=request_headers,
                    params=request_params,
                    timeout=30
                )
            else:
                # POST/PUT/DELETE等方法
                response = requests.request(
                    method=request_method,
                    url=request_url,
                    headers=request_headers,
                    json=request_json,  # 只使用json参数
                    timeout=30
                )

                # 计算执行时间
            execution_time = int((time.time() - start_time) * 1000)  # 毫秒

            # 获取响应信息
            response_status = response.status_code
            response_headers = dict(response.headers)

            try:
                response_body = response.json()
                response_body_str = json.dumps(response_body)
                global_logger.info(f'响应内容: {response_body_str}')
            except:
                response_body = response.text
                response_body_str = response_body
                global_logger.info(f'响应内容(文本): {response_body_str}')

            global_logger.info(f'HTTP请求执行完成, 状态码: {response_status}, 执行时间: {execution_time}ms')
            #HTTP请求完成后记录指标
            http_request_duration = time.time() - http_request_start_time
            # 导入metrics模块
            from utils import metrics
            metrics.record_http_request(
                method=request_method,
                status_code=response_status,
                duration_seconds=http_request_duration
            )
            # 执行断言
            is_success = True
            assertion_results = []

            if testcase['assertions']:
                global_logger.info(f'开始执行断言, 测试用例ID: {testcase_id}')
                try:
                    assertions = json.loads(testcase['assertions'])

                    for assertion in assertions:
                        assertion_type = assertion.get('type')
                        expected = assertion.get('expected')
                        actual = None
                        result = False

                        if assertion_type == 'status_code':
                            actual = response_status
                            result = str(actual) == str(expected)
                        elif assertion_type == 'json_path':
                            json_path = assertion.get('path')
                            compare_type = assertion.get('operator', assertion.get('compare_type', 'eq'))
                            if compare_type == "equals":
                                compare_type = "eq"

                            if isinstance(response_body, dict):
                                try:
                                    global_logger.info(f'执行JSON路径断言: {json_path}, 响应内容: {json.dumps(response_body)}')
                                    # 使用自定义的jsonpath_extract替代jsonpath.jsonpath
                                    matches = jsonpath_extract(response_body, json_path)
                                    if matches:
                                        actual = matches[0]

                                        global_logger.info(f'JSON路径 {json_path} 提取到值: {actual}')

                                        if compare_type == 'eq':
                                            result = str(actual) == str(expected)
                                        elif compare_type == 'contains':
                                            result = str(expected) in str(actual)
                                        elif compare_type == 'gt':
                                            result = float(actual) > float(expected)
                                        elif compare_type == 'lt':
                                            result = float(actual) < float(expected)
                                        else:
                                            result = str(actual) == str(expected)
                                    else:
                                        # 记录未找到匹配值的情况
                                        global_logger.warning(f'JSON路径 {json_path} 未找到匹配值，响应内容: {json.dumps(response_body)}')
                                        actual = None
                                except Exception as e:
                                    global_logger.error(f'执行JSON路径断言异常: {str(e)}')
                                    actual = str(e)
                        elif assertion_type == 'contains':
                            actual = response_body_str
                            result = str(expected) in str(actual)

                        assertion_result = {
                            'type': assertion_type,
                            'expected': expected,
                            'actual': actual,
                            'result': result
                        }

                        assertion_results.append(assertion_result)

                        if not result:
                            is_success = False
                        metrics.record_assertion_result(result)

                    global_logger.info(f'断言执行完成, 是否通过: {is_success}')

                    # 执行后置脚本
                    if testcase['post_script']:
                        global_logger.info(f'执行后置脚本, 测试用例ID: {testcase_id}')
                        try:
                            # 这里可以实现执行后置脚本的逻辑
                            pass
                        except Exception as e:
                            global_logger.error(f'执行后置脚本异常: {str(e)}')

                            # 保存测试结果（移到if语句外面，取消缩进）
                    current_time = int(time.time())

                    save_sql = """  
                        INSERT INTO api_result (  
                            id, testcase_id, interface_id, app_id, batch_id, request_url, request_method,  
                            request_headers, request_body, response_status, response_headers, response_body,  
                            assertion_results, is_success, error_message, execution_time, executor_id,   
                            executor_name, execute_time, test_request_id  
                        ) VALUES (  
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s  
                        )  
                    """

                    executor_id = current_user.get('id') if current_user else None
                    executor_name = current_user.get('username') if current_user else None

                    params = [
                        result_id,
                        testcase_id,
                        testcase['interface_id'],
                        testcase['app_id'],
                        None,  # batch_id为空，表示单独执行
                        request_url,
                        request_method,
                        json.dumps(request_headers),
                        json.dumps(request_data if request_data else request_json) if (
                                request_data or request_json) else None,
                        response_status,
                        json.dumps(response_headers),
                        response_body_str,
                        json.dumps(assertion_results),
                        is_success,
                        None,  # error_message
                        execution_time,
                        executor_id,
                        executor_name,
                        current_time,
                        ' '
                    ]

                    global_logger.info(f'保存测试结果，结果ID: {result_id}')
                    cursor.execute(save_sql, params)
                    conn.commit()

                    test_duration = time.time() - start_time
                    metrics.record_test_result(
                        success=is_success,
                        duration_seconds=test_duration
                    )

                    # 返回测试结果（也移到if语句外面，取消缩进）
                    return {
                        'result_id': result_id,
                        'testcase_id': testcase_id,
                        'request_url': request_url,
                        'request_method': request_method,
                        'response_status': response_status,
                        'is_success': is_success,
                        'execution_time': execution_time,
                        'execute_time': current_time
                    }
                except Exception as e:
                    global_logger.error(f'执行断言过程异常: {str(e)}')
                    is_success = False
                    assertion_results.append({
                        'type': 'error',
                        'expected': 'No error',
                        'actual': str(e),
                        'result': False
                    })



        except requests.RequestException as e:
            global_logger.error(f'HTTP请求异常: {str(e)}')

            # 计算执行时间
            execution_time = int((time.time() - start_time) * 1000)

            # 保存错误结果
            current_time = int(time.time())

            save_sql = """  
                    INSERT INTO api_result (  
                        id, testcase_id, interface_id, app_id, batch_id, request_url, request_method,  
                        request_headers, request_body, response_status, response_headers, response_body,  
                        assertion_results, is_success, error_message, execution_time, executor_id,   
                        executor_name, execute_time  
                    ) VALUES (  
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s  
                    )  
                """

            executor_id = current_user.get('id') if current_user else None
            executor_name = current_user.get('username') if current_user else None

            params = [
                result_id,
                testcase_id,
                testcase['interface_id'],
                testcase['app_id'],
                None,  # batch_id为空，表示单独执行
                request_url,
                request_method,
                json.dumps(request_headers),
                json.dumps(request_json, ensure_ascii=False) if request_json else None,
                None,  # response_status
                None,  # response_headers
                None,  # response_body
                None,  # assertion_results
                False,  # is_success
                str(e),  # error_message
                execution_time,
                executor_id,
                executor_name,
                current_time
            ]

            global_logger.info(f'保存测试结果（失败），结果ID: {result_id}')
            cursor.execute(save_sql, params)
            conn.commit()

            test_duration = time.time() - start_time
            metrics.record_test_result(
                success=False,
                duration_seconds=test_duration
            )

            # 返回测试结果
            return {
                'result_id': result_id,
                'testcase_id': testcase_id,
                'request_url': request_url,
                'request_method': request_method,
                'response_status': None,
                'is_success': False,
                'execution_time': execution_time,
                'execute_time': current_time,
                'error_message': str(e)
            }

    except Exception as e:
        global_logger.error(f'执行测试用例异常: {str(e)}')
        global_logger.error(traceback.format_exc())

        test_duration = time.time() - start_time
        metrics.record_test_result(
            success=False,
            duration_seconds=test_duration
        )
    finally:
        cursor.close()
        conn.close()

    raise e

@celery.task(bind=False)
def execute_batch_testcases_async(batch_id, testcase_ids, environment, current_user):
    """
    异步执行批量测试用例
    """
    global_logger.info(f'开始异步执行批量测试用例，批次ID: {batch_id}，用例数量: {len(testcase_ids)}')

    from utils import metrics
    metrics.update_queue_length(len(testcase_ids))
    metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='running').set(1)
    metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='completed').set(0)
    metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='failed').set(0)

    # 处理结果
    passed_cases = 0
    failed_cases = 0
    total_cases = len(testcase_ids)

    try:
        for index, testcase_id in enumerate(testcase_ids):
            try:
                # 调用执行单个测试用例的函数
                result = execute_batch_testcase(batch_id, testcase_id, environment, current_user)
                if result:
                    if result.get('is_success'):
                        passed_cases += 1
                    else:
                        failed_cases += 1
                else:
                    global_logger.error(f'执行测试用例失败，返回结果为None，用例ID: {testcase_id}')
                    failed_cases += 1

            except Exception as e:
                global_logger.error(f'执行测试用例异常，用例ID: {testcase_id}, 错误: {str(e)}')
                failed_cases += 1

                # 更新批次进度
            update_batch_progress(batch_id, passed_cases, failed_cases)

            remaining = total_cases - (index + 1)
            metrics.update_queue_length(remaining)
            completion_rate = (index + 1) / total_cases
            metrics.BATCH_COMPLETION.labels(batch_id=str(batch_id)).set(completion_rate)

        # 所有测试用例执行完成后，进行最终统计
        finalize_test_batch(batch_id)
        global_logger.info(f'异步执行批量测试用例完成，批次ID: {batch_id}，通过: {passed_cases}，失败: {failed_cases}')

        metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='running').set(0)
        metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='completed').set(1)
        success_rate = passed_cases / total_cases if total_cases > 0 else 0
        metrics.BATCH_SUCCESS_RATE.labels(batch_id=str(batch_id)).set(success_rate)

    except Exception as e:
        global_logger.error(f'异步执行批量测试用例异常: {str(e)}')
        global_logger.error(traceback.format_exc())

        # 更新批次状态为异常
        update_batch_status(batch_id, status=3)

        metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='running').set(0)
        metrics.BATCH_STATUS.labels(batch_id=str(batch_id), status='failed').set(1)

    finally:
        metrics.update_queue_length(0)



def execute_batch_testcases_sync(batch_id, testcase_ids, environment, current_user):
    """
    同步执行批量测试用例

    参数:
        batch_id: 批次ID
        testcase_ids: 测试用例ID列表
        environment: 执行环境
        current_user: 当前用户信息

    返回:
        包含所有测试结果的列表
    """
    with app.app_context():
        global_logger.info(f'开始同步执行批量测试用例，批次ID: {batch_id}，用例数量: {len(testcase_ids)}')

        results = []
        passed_cases = 0
        failed_cases = 0

        try:
            for testcase_id in testcase_ids:
                try:
                    result = execute_batch_testcase(batch_id, testcase_id, environment, current_user)
                    results.append(result)

                    if result.get('is_success'):
                        passed_cases += 1
                    else:
                        failed_cases += 1

                        # 更新批次进度
                    update_batch_progress(batch_id, passed_cases, failed_cases)

                except Exception as e:
                    global_logger.error(f'同步执行测试用例异常，用例ID: {testcase_id}, 错误: {str(e)}')
                    failed_cases += 1

                    # 更新批次进度
                    update_batch_progress(batch_id, passed_cases, failed_cases)

            # 所有测试用例执行完成后，进行最终统计
            finalize_test_batch(batch_id)
            global_logger.info(f'同步执行批量测试用例完成，批次ID: {batch_id}，通过: {passed_cases}，失败: {failed_cases}')
            return results

        except Exception as e:
            global_logger.error(f'同步执行批量测试用例异常: {str(e)}')
            global_logger.error(traceback.format_exc())

            # 更新批次状态为异常
            update_batch_status(batch_id, status=3)
            return results


def execute_batch_testcase(batch_id, testcase_id, environment, current_user):
    """
    执行批量测试中的单个测试用例

    参数:
        batch_id: 批次ID
        testcase_id: 测试用例ID
        environment: 执行环境
        current_user: 当前用户信息

    返回:
        包含执行结果的字典
    """
    global_logger.info(f'执行批量测试中的单个测试用例，批次ID: {batch_id}，用例ID: {testcase_id}')

    conn = None
    cursor = None

    try:
        # 执行测试用例
        result = execute_single_testcase(testcase_id, environment, {}, current_user)

        # 记录执行结果
        global_logger.info(f'execute_single_testcase 返回结果: {result}')
        global_logger.info(f'返回结果类型: {type(result)}')

        # 更新结果的batch_id
        conn = mysql_pool.connection()
        cursor = conn.cursor()

        if result and result.get('result_id'):  # 检查result和result_id都存在
            sql = "UPDATE api_result SET batch_id = %s WHERE id = %s"
            global_logger.info(f'执行SQL: {sql}, 参数: [{batch_id}, {result["result_id"]}]')
            cursor.execute(sql, [batch_id, result["result_id"]])
            global_logger.info(f'执行SQL完成了')
            conn.commit()
        else:
            global_logger.error(f'执行测试用例失败，无法获取结果ID，测试用例ID: {testcase_id}')
            # 即使没有result_id，也要提交事务
            conn.commit()

            # 确保返回统一格式的结果
        if result:
            return result
        else:
            # 如果execute_single_testcase返回None，构造一个失败结果
            return {
                'testcase_id': testcase_id,
                'is_success': False,
                'error_message': 'execute_single_testcase returned None',
                'result_id': None
            }

    except Exception as e:
        global_logger.error(f'执行批量测试中的单个测试用例异常，批次ID: {batch_id}，用例ID: {testcase_id}，错误: {str(e)}')
        global_logger.error(traceback.format_exc())

        # 确保数据库连接正常关闭
        if conn:
            try:
                conn.rollback()
            except:
                pass

                # 返回失败结果
        return {
            'testcase_id': testcase_id,
            'is_success': False,
            'error_message': str(e),
            'result_id': None
        }

    finally:
        # 确保数据库连接被正确关闭
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if conn:
            try:
                conn.close()
            except:
                pass

def create_test_batch(batch_id, app_id, test_request_id, name, total_cases, current_user, ai_generated=0):
    """
    创建测试批次

    参数:
        batch_id: 批次ID
        app_id: 应用ID
        test_request_id: 测试需求ID（可选）
        name: 批次名称
        total_cases: 总用例数
        current_user: 当前用户信息
        ai_generated: 是否AI生成（0-否，1-是）
    """
    global_logger.info(f'创建测试批次，ID: {batch_id}，名称: {name}，应用ID: {app_id}，AI生成: {ai_generated}')

    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        current_time = int(time.time())

        executor_id = current_user.get('id') if current_user else None
        executor_name = current_user.get('username') if current_user else None

        #  添加ai_generated字段
        sql = """  
                INSERT INTO api_test_batch (  
                    id, name, app_id, test_request_id, total_cases, passed_cases,  
                    failed_cases, status, trigger_type, executor_id, executor_name,  
                    create_time, ai_generated  
                ) VALUES (  
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s  
                )  
            """

        params = [
            batch_id,
            name,
            app_id,
            test_request_id,
            total_cases,
            0,  # passed_cases
            0,  # failed_cases
            1,  # status: 1-执行中
            1,  # trigger_type: 1-手动触发
            executor_id,
            executor_name,
            current_time,
            ai_generated  #  AI生成标记
        ]

        global_logger.info(f'执行SQL: {sql}, 参数: {params}')
        cursor.execute(sql, params)
        conn.commit()

        global_logger.info(f'创建测试批次成功，ID: {batch_id}')

    except Exception as e:
        global_logger.error(f'创建测试批次异常，ID: {batch_id}，错误: {str(e)}')
        global_logger.error(traceback.format_exc())
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()


def update_batch_progress(batch_id, passed_cases, failed_cases):
    """
    更新测试批次进度

    参数:
        batch_id: 批次ID
        passed_cases: 通过用例数
        failed_cases: 失败用例数
    """
    global_logger.info(f'更新测试批次进度，ID: {batch_id}，通过: {passed_cases}，失败: {failed_cases}')

    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        sql = """  
                UPDATE api_test_batch  
                SET passed_cases = %s, failed_cases = %s  
                WHERE id = %s  
            """

        params = [passed_cases, failed_cases, batch_id]

        global_logger.info(f'执行SQL: {sql}, 参数: {params}')
        cursor.execute(sql, params)
        conn.commit()

        global_logger.info(f'更新测试批次进度成功，ID: {batch_id}')

    except Exception as e:
        global_logger.error(f'更新测试批次进度异常，ID: {batch_id}，错误: {str(e)}')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()


def update_batch_status(batch_id, status=2):
    """
    更新测试批次状态

    参数:
        batch_id: 批次ID
        status: 状态，1-执行中，2-已完成，3-异常
    """
    global_logger.info(f'更新测试批次状态，ID: {batch_id}，状态: {status}')

    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        current_time = int(time.time())

        sql = """  
            UPDATE api_test_batch  
            SET status = %s, end_time = %s  
            WHERE id = %s  
        """

        params = [status, current_time, batch_id]

        global_logger.info(f'执行SQL: {sql}, 参数: {params}')
        cursor.execute(sql, params)
        conn.commit()

        global_logger.info(f'更新测试批次状态成功，ID: {batch_id}')

    except Exception as e:
        global_logger.error(f'更新测试批次状态异常，ID: {batch_id}，错误: {str(e)}')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()


def generate_excel_report(batch_id):
    """
    生成Excel测试报告

    参数:
        batch_id: 批次ID

    返回:
        Excel文件的二进制数据
    """
    global_logger.info(f'开始生成Excel测试报告，批次ID: {batch_id}')

    # 获取批次信息和测试结果
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 查询批次信息
        batch_sql = """  
            SELECT name, app_id, total_cases, passed_cases, failed_cases,   
                   executor_name, create_time, end_time, status  
            FROM api_test_batch  
            WHERE id = %s  
        """
        global_logger.info(f'执行SQL: {batch_sql}, 参数: [{batch_id}]')
        cursor.execute(batch_sql, [batch_id])
        batch = cursor.fetchone()

        if not batch:
            global_logger.error(f'未找到测试批次，ID: {batch_id}')
            raise Exception(f"未找到测试批次: {batch_id}")

            # 查询应用信息
        app_sql = "SELECT name FROM api_application WHERE id = %s"
        global_logger.info(f'执行SQL: {app_sql}, 参数: [{batch["app_id"]}]')
        cursor.execute(app_sql, [batch["app_id"]])
        app_info = cursor.fetchone()

        # 查询测试结果
        result_sql = """  
            SELECT r.id, r.testcase_id, t.name as testcase_name,   
                   r.request_url, r.request_method, r.response_status,   
                   r.is_success, r.execution_time, r.execute_time,  
                   r.error_message  
            FROM api_result r  
            LEFT JOIN api_testcase t ON r.testcase_id = t.id  
            WHERE r.batch_id = %s  
            ORDER BY r.execute_time  
        """
        global_logger.info(f'执行SQL: {result_sql}, 参数: [{batch_id}]')
        cursor.execute(result_sql, [batch_id])
        results = cursor.fetchall()

        global_logger.info(f'获取测试数据成功，批次: {batch["name"]}, 结果数量: {len(results)}')

        # 创建Excel文件
        workbook = openpyxl.Workbook()

        # 创建概览工作表
        overview_sheet = workbook.active
        overview_sheet.title = "测试概览"

        # 设置概览标题
        overview_sheet['A1'] = "测试报告概览"
        overview_sheet['A1'].font = Font(size=14, bold=True)
        overview_sheet.merge_cells('A1:G1')

        # 设置概览信息
        overview_sheet['A3'] = "批次名称:"
        overview_sheet['B3'] = batch['name']
        overview_sheet['A4'] = "应用名称:"
        overview_sheet['B4'] = app_info['name'] if app_info else ""
        overview_sheet['A5'] = "执行人:"
        overview_sheet['B5'] = batch['executor_name'] or ""

        overview_sheet['D3'] = "开始时间:"
        overview_sheet['E3'] = datetime.fromtimestamp(batch['create_time']).strftime('%Y-%m-%d %H:%M:%S') if batch[
            'create_time'] else ""
        overview_sheet['D4'] = "结束时间:"
        overview_sheet['E4'] = datetime.fromtimestamp(batch['end_time']).strftime('%Y-%m-%d %H:%M:%S') if batch[
            'end_time'] else ""
        overview_sheet['D5'] = "执行状态:"
        status_map = {1: "执行中", 2: "已完成", 3: "异常"}
        overview_sheet['E5'] = status_map.get(batch['status'], "未知")

        # 设置统计信息
        overview_sheet['A7'] = "测试统计"
        overview_sheet['A7'].font = Font(size=12, bold=True)
        overview_sheet.merge_cells('A7:G7')

        overview_sheet['A9'] = "总用例数:"
        overview_sheet['B9'] = batch['total_cases'] or 0
        overview_sheet['A10'] = "通过用例数:"
        overview_sheet['B10'] = batch['passed_cases'] or 0
        overview_sheet['A11'] = "失败用例数:"
        overview_sheet['B11'] = batch['failed_cases'] or 0

        overview_sheet['D9'] = "通过率:"
        pass_rate = round((batch['passed_cases'] or 0) * 100 / (batch['total_cases'] or 1), 2)
        overview_sheet['E9'] = f"{pass_rate}%"

        # 创建详细结果工作表
        detail_sheet = workbook.create_sheet(title="测试详情")

        # 设置表头
        headers = ["序号", "测试用例", "请求URL", "请求方法", "响应状态", "是否通过", "执行时间(ms)", "执行时间",
                   "错误信息"]
        for col, header in enumerate(headers, 1):
            detail_sheet.cell(row=1, column=col, value=header).font = Font(bold=True)

            # 填充数据
        for row, result in enumerate(results, 2):
            detail_sheet.cell(row=row, column=1, value=row - 1)
            detail_sheet.cell(row=row, column=2, value=result['testcase_name'] or f"用例 {result['testcase_id']}")
            detail_sheet.cell(row=row, column=3, value=result['request_url'])
            detail_sheet.cell(row=row, column=4, value=result['request_method'])
            detail_sheet.cell(row=row, column=5, value=result['response_status'])
            detail_sheet.cell(row=row, column=6, value="通过" if result['is_success'] else "失败")
            detail_sheet.cell(row=row, column=7, value=result['execution_time'])

            execute_time_str = ""
            if result['execute_time']:
                execute_time_str = datetime.fromtimestamp(result['execute_time']).strftime('%Y-%m-%d %H:%M:%S')
            detail_sheet.cell(row=row, column=8, value=execute_time_str)

            detail_sheet.cell(row=row, column=9, value=result['error_message'] or "")

            # 调整列宽
        for sheet in [overview_sheet, detail_sheet]:
            for col in sheet.columns:
                max_length = 0
                from openpyxl.utils import get_column_letter

                try:
                    if hasattr(col[0], 'column_letter'):
                        column = col[0].column_letter
                    elif hasattr(col[0], 'column'):
                        column = get_column_letter(col[0].column)
                    else:
                        # 默认值
                        column = get_column_letter(1)  # 'A'
                except Exception as e:
                    # 记录错误但继续执行
                    print(f"获取列字母时出错: {str(e)}")
                    column = get_column_letter(1)  # 默认使用'A'

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = min(adjusted_width, 50)

                # 保存到内存
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        global_logger.info(f'生成Excel测试报告成功，批次ID: {batch_id}')

        return excel_buffer.getvalue()

    except Exception as e:
        global_logger.error(f'生成Excel测试报告异常: {str(e)}')
        global_logger.error(traceback.format_exc())
        raise e
    finally:
        cursor.close()
        conn.close()


# 简化PDF生成功能，不使用ReportLab
def generate_pdf_report(batch_id):
    """
    生成PDF测试报告 (简化版，实际返回HTML内容)

    参数:
        batch_id: 批次ID

    返回:
        HTML格式的报告
    """
    global_logger.info(f'开始生成HTML测试报告，批次ID: {batch_id}')

    # 获取批次信息和测试结果
    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
        # 查询批次信息
        batch_sql = """  
            SELECT name, app_id, total_cases, passed_cases, failed_cases,   
                   executor_name, create_time, end_time, status  
            FROM api_test_batch  
            WHERE id = %s  
        """
        global_logger.info(f'执行SQL: {batch_sql}, 参数: [{batch_id}]')
        cursor.execute(batch_sql, [batch_id])
        batch = cursor.fetchone()

        if not batch:
            global_logger.error(f'未找到测试批次，ID: {batch_id}')
            raise Exception(f"未找到测试批次: {batch_id}")

            # 查询应用信息
        app_sql = "SELECT name FROM api_application WHERE id = %s"
        global_logger.info(f'执行SQL: {app_sql}, 参数: [{batch["app_id"]}]')
        cursor.execute(app_sql, [batch["app_id"]])
        app_info = cursor.fetchone()

        # 查询测试结果
        result_sql = """  
            SELECT r.id, r.testcase_id, t.name as testcase_name,   
                   r.request_url, r.request_method, r.response_status,   
                   r.is_success, r.execution_time, r.execute_time,  
                   r.error_message  
            FROM api_result r  
            LEFT JOIN api_testcase t ON r.testcase_id = t.id  
            WHERE r.batch_id = %s  
            ORDER BY r.execute_time  
        """
        global_logger.info(f'执行SQL: {result_sql}, 参数: [{batch_id}]')
        cursor.execute(result_sql, [batch_id])
        results = cursor.fetchall()

        global_logger.info(f'获取测试数据成功，批次: {batch["name"]}, 结果数量: {len(results)}')

        # 生成HTML报告
        html = f"""  
        <!DOCTYPE html>  
        <html>  
        <head>  
            <meta charset="UTF-8">  
            <title>API测试报告 - {batch['name']}</title>  
            <style>  
                body {{ font-family: Arial, sans-serif; margin: 20px; }}  
                h1, h2 {{ color: #333; }}  
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}  
                th, td {{ border: 1px solid #ddd; padding: 8px; }}  
                th {{ background-color: #f2f2f2; text-align: left; }}  
                tr:nth-child(even) {{ background-color: #f9f9f9; }}  
                .success {{ color: green; }}  
                .failure {{ color: red; }}  
            </style>  
        </head>  
        <body>  
            <h1>API自动化测试报告</h1>  

            <h2>批次信息</h2>  
            <table>  
                <tr>  
                    <th>批次名称</th>  
                    <td>{batch['name']}</td>  
                    <th>应用名称</th>  
                    <td>{app_info['name'] if app_info else ""}</td>  
                </tr>  
                <tr>  
                    <th>执行人</th>  
                    <td>{batch['executor_name'] or ""}</td>  
                    <th>执行状态</th>  
                    <td>{{1: "执行中", 2: "已完成", 3: "异常"}}.get(batch['status'], "未知")</td>  
                </tr>  
                <tr>  
                    <th>开始时间</th>  
                    <td>{datetime.fromtimestamp(batch['create_time']).strftime('%Y-%m-%d %H:%M:%S') if batch['create_time'] else ""}</td>  
                    <th>结束时间</th>  
                    <td>{datetime.fromtimestamp(batch['end_time']).strftime('%Y-%m-%d %H:%M:%S') if batch['end_time'] else ""}</td>  
                </tr>  
            </table>  

            <h2>测试统计</h2>  
            <table>  
                <tr>  
                    <th>总用例数</th>  
                    <td>{batch['total_cases'] or 0}</td>  
                    <th>通过用例数</th>  
                    <td>{batch['passed_cases'] or 0}</td>  
                </tr>  
                <tr>  
                    <th>失败用例数</th>  
                    <td>{batch['failed_cases'] or 0}</td>  
                    <th>通过率</th>  
                    <td>{round((batch['passed_cases'] or 0) * 100 / (batch['total_cases'] or 1), 2)}%</td>  
                </tr>  
            </table>  

            <h2>测试详情</h2>  
            <table>  
                <tr>  
                    <th>序号</th>  
                    <th>测试用例</th>  
                    <th>请求URL</th>  
                    <th>请求方法</th>  
                    <th>响应状态</th>  
                    <th>是否通过</th>  
                    <th>执行时间(ms)</th>  
                    <th>执行时间</th>  
                    <th>错误信息</th>  
                </tr>  
        """

        # 添加测试结果数据
        for i, result in enumerate(results, 1):
            execute_time_str = ""
            if result['execute_time']:
                execute_time_str = datetime.fromtimestamp(result['execute_time']).strftime('%Y-%m-%d %H:%M:%S')

            result_class = "success" if result['is_success'] else "failure"
            result_text = "通过" if result['is_success'] else "失败"

            html += f"""  
                <tr>  
                    <td>{i}</td>  
                    <td>{result['testcase_name'] or f"用例 {result['testcase_id']}"}</td>  
                    <td>{result['request_url']}</td>  
                    <td>{result['request_method']}</td>  
                    <td>{result['response_status'] or "-"}</td>  
                    <td class="{result_class}">{result_text}</td>  
                    <td>{result['execution_time'] or "-"}</td>  
                    <td>{execute_time_str}</td>  
                    <td>{result['error_message'] or ""}</td>  
                </tr>  
            """

        html += """  
            </table>  
        </body>  
        </html>  
        """

        global_logger.info(f'生成HTML测试报告成功，批次ID: {batch_id}')

        return html.encode('utf-8')

    except Exception as e:
        global_logger.error(f'生成HTML测试报告异常: {str(e)}')
        global_logger.error(traceback.format_exc())
        raise e
    finally:
        cursor.close()
        conn.close()

    # 定义一个简化版的get_current_user函数
def get_current_user():
    """
                获取当前用户信息

                返回:
                    包含用户信息的字典
                """
    # 这里应该实现从请求中获取用户信息的逻辑
    # 简单起见，返回一个默认用户
    return {
        'id': 1,
        'username': 'admin'
    }

# 替代jsonpath库
def jsonpath_extract(obj, path):
    """
    简单实现jsonpath提取功能

    参数:
        obj: JSON对象
        path: 路径表达式，例如 $.data.items[0].name

    返回:
        匹配的值列表
    """
    global_logger.info(f'使用自定义jsonpath提取: {path}')

    # 去掉开头的$
    if path.startswith('$'):
        path = path[1:]

        # 如果路径以.开头，去掉开头的.
    if path.startswith('.'):
        path = path[1:]

        # 按.分割路径
    parts = path.split('.')

    # 过滤掉空字符串
    parts = [part for part in parts if part]

    current = obj
    try:
        for part in parts:
            # 处理数组索引，如items[0]
            if '[' in part and ']' in part:
                array_name = part.split('[')[0]
                index_str = part.split('[')[1].split(']')[0]

                if array_name:
                    current = current[array_name]

                    # 处理索引
                index = int(index_str)
                current = current[index]
            else:
                current = current[part]

        return [current]
    except (KeyError, IndexError, TypeError) as e:
        global_logger.warning(f'jsonpath提取失败: {str(e)}')
        return []
def finalize_test_batch(batch_id):
    """
    完成测试批次，统计结果并更新状态
    参数:
         batch_id: 批次ID
    """
    global_logger.info(f'完成测试批次，ID: {batch_id}')

    conn = mysql_pool.connection()
    cursor = conn.cursor()

    try:
         # 统计测试结果
        stats_sql = """  
            SELECT   
                COUNT(*) as total,  
                SUM(CASE WHEN is_success = 1 THEN 1 ELSE 0 END) as passed,  
                SUM(CASE WHEN is_success = 0 THEN 1 ELSE 0 END) as failed  
            FROM api_result  
            WHERE batch_id = %s  
        """

        cursor.execute(stats_sql, [batch_id])
        stats = cursor.fetchone()

        if not stats:
            global_logger.warning(f'未找到测试批次的结果数据，ID: {batch_id}')
            return

        passed = stats['passed'] or 0
        failed = stats['failed'] or 0

        global_logger.info(f'测试批次最终统计，ID: {batch_id}，通过: {passed}，失败: {failed}')

        # 更新批次进度
        update_batch_progress(batch_id, passed, failed)

        # 更新批次状态为已完成
        update_batch_status(batch_id, 2)  # 2-已完成

    except Exception as e:
        global_logger.error(f'完成测试批次异常，ID: {batch_id}，错误: {str(e)}')
        global_logger.error(traceback.format_exc())
    finally:
        cursor.close()
        conn.close()


        # 启动后台任务线程池
#batch_update_lock = threading.Lock()  # 用于批次更新的线程锁
def verify_batch_exists(batch_id):
    """验证批次是否存在"""
    try:
        conn = mysql_pool.connection()
        cursor = conn.cursor()

        sql = "SELECT id FROM api_test_batch WHERE id = %s"
        cursor.execute(sql, [batch_id])
        result = cursor.fetchone()

        cursor.close()
        conn.close()

        return result is not None
    except Exception as e:
        global_logger.error(f"验证批次存在性失败: {str(e)}")
        return False


def update_batch_status_to_running(batch_id):
    """更新批次状态为执行中"""
    try:
        conn = mysql_pool.connection()
        cursor = conn.cursor()

        sql = "UPDATE api_test_batch SET status = 2 WHERE id = %s"  # 2=执行中
        cursor.execute(sql, [batch_id])
        conn.commit()

        cursor.close()
        conn.close()

        global_logger.info(f"批次 {batch_id} 状态更新为执行中")
    except Exception as e:
        global_logger.error(f"更新批次状态失败: {str(e)}")